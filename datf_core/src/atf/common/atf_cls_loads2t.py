from pyspark.sql import SparkSession
from pyspark.sql.functions import * 
from pyspark.sql.types import *
from pyspark.sql.types import StructType,StructField, StringType, IntegerType,DoubleType,DateType
from datetime import datetime
from sys import *
import time
import pandas as pd
from openpyxl import load_workbook
from constants import *
from IPython.display import display
from atf.common.atf_common_functions import log_info

class LoadS2T:
   
  def __init__(self, configFilePath, spark):
    self.stageEnabled = True
    self.spark = spark
    config_wb = load_workbook(configFilePath, read_only=True)
    if 'StageMapping' not in config_wb.sheetnames:
      self.stageEnabled = False
    
    
    config_df = pd.read_excel(configFilePath, engine='openpyxl',sheet_name='MappingConfiguration', header = None, keep_default_na=False)
    config = {}
    for i in range(len(config_df)):
      config[config_df[0][i]] = config_df[1][i]
      
    self.mappingName=config["mappingname"]
    self.mappingExtendedProperties1=config["mappingextendedproperties1"]
    self.mappingExtendedProperties2=config["mappingextendedproperties2"]
    self.mappingExtendedProperties3=config["mappingextendedproperties3"]    

    self.sourceConnectionName=config["sourceconnectionname"]
    self.sourceConnectionType=config["sourceconnectiontype"]
    #self.sourceType=config["sourcedatatype"]
    self.sourceFileFormat=config["sourcefileformat"]

    if config["sourcefilepath"] != "":
      #self.sourceconnectionval = get_connection_config(self.sourceConnectionName)['BUCKETNAME']
      #self.sourceFilePath=get_mount_path(self.sourceconnectionval + config["sourcefilepath"])
      if 'Volumes' in config["sourcefilepath"]:
        self.sourceFilePath = config["sourcefilepath"]
      elif 'abfs' in config["sourcefilepath"]:
        self.sourceFilePath = config["sourcefilepath"]
      else:
        log_info("Appending Root path to relative file Path")
        self.sourceFilePath= root_path+config["sourcefilepath"]
        log_info(f"Source File Path :- {self.sourceFilePath}")

    self.sourceFileName=config["sourcefilename"]
    if config["sourcefilehasheader"].upper() == 'Y':
      self.sourceFileHasHeader=True
    else:
      self.sourceFileHasHeader=False    
    self.sourceFileDelimiter=config["sourcefiledelimiter"]
    self.sourceTimestampFormat=config["sourcetimestampformat"]    
    self.sourceDatabaseSchemaName=config["sourcedatabaseschemaname"]
    self.sourceDatabaseTableName=config["sourcedatabasetablename"]
    
    self.stageAliasName=config["stagealiasname"]
    self.stageConnectionName=config["stageconnectionname"]
    self.stageConnectionType=config["stageconnectiontype"]
    #self.stageDataType=config["stagedatatype"]
    self.stageFileFormat=config["stagefileformat"]
    self.stageFileDescription=config["stagefiledescription"]
    self.stageFileName=config["stagefilename"]
    if config["stagefilepath"] != "":
      #self.stageconnectionval = get_connection_config(self.stageConnectionName)['BUCKETNAME']
      #self.stageFilePath=get_mount_path(self.stageconnectionval +config["stagefilepath"])
      if 'Volumes' in config["stagefilepath"]:
        self.stageFilePath = config["stagefilepath"]
      elif 'abfs' in config["stagefilepath"]:
        self.stageFilePath = config["stagefilepath"]
      else:
        log_info("Appending Root path to relative file Path")
        self.stageFilePath=root_path+config["stagefilepath"]
        log_info(f"Stage File Path :- {self.stageFilePath}")

    self.stageFileHasHeader=config["stagefilehasheader"]
    self.stageFileDelimiter=config["stagefiledelimiter"]
    self.stageTimestampFormat=config["stagetimestampformat"]
    self.stageDatabaseSchemaName=config["stagedatabaseschemaname"]
    self.stageDatabaseTableName=config["stagedatabasetablename"]

    self.targetAliasName=config["targetaliasname"]
    self.targetConnectionName=config["targetconnectionname"]
    self.targetConnectionType=config["targetconnectiontype"]
    #self.targetDataType=config["targetdatatype"]
    self.targetFileFormat=config["targetfileformat"]
    self.targetFileDescription=config["targetfiledescription"]
    self.targetFileName=config["targetfilename"]
    if config["targetfilepath"] != "":
      #self.targetconnectionval = get_connection_config(self.targetConnectionName)['BUCKETNAME']
      #self.targetFilePath=get_mount_path(self.targetconnectionval +config["targetfilepath"])
      if 'Volumes' in config["targetfilepath"]:
        self.targetFilePath = config["targetfilepath"]
      elif 'abfs' in config["targetfilepath"]:
        self.targetFilePath = config["targetfilepath"]
      else:
        log_info("Appending Root path to relative file Path")
        self.targetFilePath=root_path+config["targetfilepath"]
        log_info(f"Target File Path :- {self.targetFilePath}")

    self.targetFileHasHeader=config["targetfilehasheader"]
    self.targetFileDelimiter=config["targetfiledelimiter"]
    self.targetTimestampFormat=config["targettimestampformat"]
    self.targetDatabaseSchemaName=config["targetdatabaseschemaname"]
    self.targetDatabaseTableName=config["targetdatabasetablename"]    
    
    if config["trimwhitespaces"].upper() == 'Y':
      self.trimWhiteSpaces=True
    else:
      self.trimWhiteSpaces=False
    #self.convertCase=config["convertcase"].upper()
    if config["removeduplicates"].upper() == 'Y':
      self.removeDuplicates=True
    else:
      self.removeDuplicates=False

    if config["enableblacklisting"].upper() == 'Y':
      self.blacklistingEnabled=True
    else:
      self.blacklistingEnabled=False 
      
    self.stageLoadStrategy = config["stageloadstrategy"]
    if config["truncatestagebeforeload"].upper() == 'Y':
      self.truncateStageBeforeLoad=True
    else:
      self.truncateStageBeforeLoad=False
    self.targetLoadStrategy = config["targetloadstrategy"]
    if config["truncatetargetbeforeload"].upper() == 'Y':
      self.truncateTargetBeforeLoad=True
    else:
      self.truncateTargetBeforeLoad=False      

    self.preProcessorFunction1=config["preprocessorfunction1"]
    self.preProcessorFunction2=config["preprocessorfunction2"]
    self.preProcessorFunction3=config["preprocessorfunction3"]

    self.preProcessorLst=[]
    if self.preProcessorFunction1 != "":
      self.preProcessorLst.append(self.preProcessorFunction1)
    if self.preProcessorFunction2 != "":
      self.preProcessorLst.append(self.preProcessorFunction2)
    if self.preProcessorFunction3 != "":
      self.preProcessorLst.append(self.preProcessorFunction3)  
    
    if len(self.preProcessorLst) > 0:
      self.preProcessEnabled = True
    else:
      self.preProcessEnabled = False
      
    self.sourceTableName=""
    self.stageTableName=""
    self.targetTableName=""
    if self.sourceFileName != "":
      self.sourceFile=self.sourceFilePath+'/'+self.sourceFileName
      self.sourceTableName=f"{self.sourceFileFormat}.`{self.sourceFile}`"
    elif self.sourceDatabaseTableName != "":
      self.sourceTableName=f"{self.sourceDatabaseSchemaName}.{self.sourceDatabaseTableName}"
      self.sourceFile = self.sourceTableName
      
    if self.stageFileName != "":
      self.stageFile=self.stageFilePath+'/'+self.stageFileName
      self.stageTableName=f"{self.stageFileFormat}.`{self.stageFile}`"
    elif self.stageDatabaseTableName != "":
      self.stageTableName=f"{self.stageDatabaseSchemaName}.{self.stageDatabaseTableName}"
      self.stageFile = self.stageTableName
      
    if self.targetFileName != "":
      self.targetFile=self.targetFilePath+'/'+self.targetFileName
      self.targetTableName=f"{self.targetFileFormat}.`{self.targetFile}`"
    elif self.targetDatabaseTableName != "":
      self.targetTableName=f"{self.targetDatabaseSchemaName}.{self.targetDatabaseTableName}"
      self.targetFile = self.targetTableName
   
    self.schema_pddf=pd.read_excel(configFilePath, engine='openpyxl',sheet_name='Schema')
    self.schema_pddf=self.schema_pddf.fillna("")
    print(type(spark))
    self.schema_df=spark.createDataFrame(self.schema_pddf)
      
    self.sourceschema_df=self.schema_df.filter(col("tabletype") == 'source')
    self.stageschema_df=self.schema_df.filter(col("tabletype") == 'stage')
    self.targetschema_df=self.schema_df.filter(col("tabletype") == 'target') 
   
    self.stagemapping_df=spark.createDataFrame([], StructType([])) #Added by Susan

    if self.stageEnabled == True:
      self.stagemapping_pddf=pd.read_excel(configFilePath, engine='openpyxl',sheet_name='StageMapping', keep_default_na=False)
      self.stagemapping_df=spark.createDataFrame(self.stagemapping_pddf)
      
    self.targetmapping_pddf=pd.read_excel(configFilePath, engine='openpyxl',sheet_name='TargetMapping', keep_default_na=False)
    self.targetmapping_df=spark.createDataFrame(self.targetmapping_pddf)
  
    self.tgtschema_df = (self.targetschema_df
                         .select("columnname","datatype","length","scale")
                         .withColumnRenamed("datatype","targetcolumndatatype")
                         .withColumnRenamed("columnname","targetcolumnname")
                         .withColumnRenamed("length","targetlength")
                         .withColumnRenamed("scale","targetscale"))
    self.stgschema_df = (self.stageschema_df
                         .select("columnname","datatype","length","scale")
                         .withColumnRenamed("datatype","stagecolumndatatype")
                         .withColumnRenamed("columnname","stagecolumnname")
                         .withColumnRenamed("length","stagelength")
                         .withColumnRenamed("scale","stagescale"))
    self.srcschema_df = (self.sourceschema_df
                         .select("columnname","datatype","length","scale")
                         .withColumnRenamed("datatype","sourcecolumndatatype")
                         .withColumnRenamed("columnname","sourcecolumnname")
                         .withColumnRenamed("length","sourcelength")
                         .withColumnRenamed("scale","sourcescale"))
 
    if self.stageEnabled == True:
      #StageMapping DF with new joined fields stagecolumndatatype and sourcecolumndatatype    
      self.stageschemamap_df = (self.stagemapping_df
                               .join(self.stgschema_df, self.stagemapping_df.tgtcolumnname == self.stgschema_df.stagecolumnname,"left_outer")
                               .join(self.srcschema_df, self.stagemapping_df.srccolumnname == self.srcschema_df.sourcecolumnname,"left_outer")
                               .drop("stagecolumnname","sourcecolumnname"))   
      #TargetMapping DF with new joined fields targetcolumndatatype and stagecolumndatatype
      self.targetschemamap_df = (self.targetmapping_df
                               .join(self.tgtschema_df, self.targetmapping_df.tgtcolumnname == self.tgtschema_df.targetcolumnname,"left_outer")
                               .join(self.stgschema_df, self.targetmapping_df.srccolumnname == self.stgschema_df.stagecolumnname,"left_outer")
                               .drop("targetcolumnname","stagecolumnname"))
    elif self.stageEnabled == False:
      #TargetMapping DF with new joined fields targetcolumndatatype and sourcecolumndatatype
      self.targetschemamap_df = (self.targetmapping_df
                               .join(self.tgtschema_df, self.targetmapping_df.tgtcolumnname == self.tgtschema_df.targetcolumnname,"left_outer")
                               .join(self.srcschema_df, self.targetmapping_df.srccolumnname == self.srcschema_df.sourcecolumnname,"left_outer")
                               .drop("targetcolumnname","sourcecolumnname"))
      
      
  def getSchemaStruct(self,entitytype):
      schemaStruct = StructType()
      if entitytype == "source":
        schemadfCollect = self.sourceschema_df.collect()
      elif entitytype == "target":
        schemadfCollect = self.targetschema_df.collect()
      for row in schemadfCollect:
        #print(row['sourcecolumnname'] + "," +row['sourcecolumndatatype'])
        if row['datatype'] == "string" :
          schemaStruct.add(StructField(row['columnname'],StringType(),True))
        elif row['datatype'] == "datetime" :
          schemaStruct.add(StructField(row['columnname'],DateType(),True))
        elif row['datatype'] == "double" :
          schemaStruct.add(StructField(row['columnname'],DoubleType(),True))
        elif row['datatype'] == "integer" :
          schemaStruct.add(StructField(row['columnname'],IntegerType(),True))
      return schemaStruct
